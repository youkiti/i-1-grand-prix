import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Iterator, Optional
from collections import defaultdict
from dataclasses import dataclass

MESSAGE_CANDIDATE_COLUMNS = ["message", "content", "text", "body"]


@dataclass
class DocumentWithMetadata:
    """メタデータ付きドキュメント"""
    filename: str
    content: str
    url: str = ""
    page_title: str = ""
    link_text: str = ""
    page_url: str = ""
    date: str = ""  # 審議会の開催日付など

    def to_enriched_content(self) -> str:
        """メタデータヘッダー付きのコンテンツを返す"""
        header_lines = ["[Document Info]"]
        header_lines.append(f"- file: {self.filename}")
        if self.url:
            header_lines.append(f"- url: {self.url}")
        if self.date:
            header_lines.append(f"- date: {self.date}")
        if self.page_title:
            header_lines.append(f"- page_title: {self.page_title}")
        if self.link_text:
            header_lines.append(f"- link_text: {self.link_text}")
        if self.page_url:
            header_lines.append(f"- page_url: {self.page_url}")
        header_lines.append("[/Document Info]")
        header_lines.append("")

        return "\n".join(header_lines) + self.content


def extract_date_from_metadata(page_title: str, link_text: str, filename: str) -> str:
    """
    メタデータから日付を抽出する
    
    優先順位:
    1. page_title から抽出（例: 「第１６回会議（令和６年８月２４日開催）」）
    2. link_text から抽出
    3. filename から抽出（例: 20210414diji01_minutes.pdf）
    
    Returns:
        抽出された日付文字列（見つからない場合は空文字）
    """
    # 和暦パターン: 令和X年Y月Z日、平成X年Y月Z日 等
    wareki_pattern = r'[（(]?[令平昭]和[０-９0-9元]+年[０-９0-9]+月[０-９0-9]+日[開催）)]*'
    
    # より厳密な和暦パターン（開催日を含む）
    wareki_full_pattern = r'(?:令和|平成|昭和)[０-９0-9元]+年[０-９0-9]+月[０-９0-9]+日(?:開催)?'
    
    # 西暦パターン: 2021年4月14日、2021-04-14、2021/04/14
    seireki_pattern = r'20[0-9]{2}[-/年][0-9]{1,2}[-/月][0-9]{1,2}日?'
    
    # ファイル名からの日付パターン: 20210414 (YYYYMMDDの8桁)
    filename_date_pattern = r'^(20[0-9]{2})(0[1-9]|1[0-2])(0[1-9]|[12][0-9]|3[01])'
    
    def normalize_japanese_numerals(text: str) -> str:
        """全角数字を半角に変換"""
        trans_table = str.maketrans('０１２３４５６７８９', '0123456789')
        return text.translate(trans_table)
    
    def format_wareki_date(match_text: str) -> str:
        """和暦マッチを整形して返す"""
        # 括弧や「開催」を除去して日付部分だけ返す
        cleaned = re.sub(r'[（()）開催]', '', match_text)
        return normalize_japanese_numerals(cleaned)
    
    # 1. page_title から検索
    if page_title:
        match = re.search(wareki_full_pattern, page_title)
        if match:
            return format_wareki_date(match.group())
        match = re.search(seireki_pattern, page_title)
        if match:
            return match.group()
    
    # 2. link_text から検索
    if link_text:
        match = re.search(wareki_full_pattern, link_text)
        if match:
            return format_wareki_date(match.group())
        match = re.search(seireki_pattern, link_text)
        if match:
            return match.group()
    
    # 3. filename から検索
    if filename:
        match = re.match(filename_date_pattern, filename)
        if match:
            year, month, day = match.groups()
            return f"{year}-{month}-{day}"
    
    return ""


def load_scraper_metadata(folder_path: Path) -> Dict[str, dict]:
    """
    スクレイパーが生成したメタデータファイルを読み込む
    
    metadata.json または metadata_*.json パターンのファイルを読み込む。
    複数ファイルが存在する場合はマージする。

    Returns:
        {filename: metadata_dict} の辞書
    """
    combined_files = {}
    
    # metadata.json を読み込み
    metadata_path = folder_path / "metadata.json"
    if metadata_path.exists():
        with open(metadata_path, "r", encoding="utf-8") as f:
            data = json.load(f)
            combined_files.update(data.get("files", {}))
    
    # metadata_*.json パターンのファイルを読み込み
    for meta_file in folder_path.glob("metadata_*.json"):
        try:
            with open(meta_file, "r", encoding="utf-8") as f:
                data = json.load(f)
                combined_files.update(data.get("files", {}))
        except (json.JSONDecodeError, IOError) as e:
            print(f"Warning: Failed to read {meta_file}: {e}")
    
    return combined_files


def load_messages_csv(path: Path) -> List[Dict[str, str]]:
    """CSVをリストの辞書として読み込む（pandas不使用）"""
    if not path.exists():
        raise FileNotFoundError(f"CSVが見つかりません: {path}")

    rows = []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            rows.append(row)

    if not rows:
        raise ValueError("CSVにデータがありません")

    if "session_id" not in rows[0]:
        raise ValueError("CSVに session_id 列が必要です")

    return rows


def pick_message_column(rows: List[Dict[str, str]]) -> str:
    """メッセージ列を特定"""
    if not rows:
        raise ValueError("データが空です")

    columns = rows[0].keys()
    for col in MESSAGE_CANDIDATE_COLUMNS:
        if col in columns:
            return col
    raise ValueError(f"メッセージ列が見つかりません。候補: {', '.join(MESSAGE_CANDIDATE_COLUMNS)}")


def build_sessions_data(rows: List[Dict[str, str]]) -> Tuple[List[str], str, Dict[str, int]]:
    """セッションごとにデータをグループ化"""
    message_col = pick_message_column(rows)

    # session_id でグループ化
    sessions_dict = defaultdict(list)
    for row in rows:
        session_id = row.get("session_id", "")
        sessions_dict[session_id].append(row)

    sessions: List[str] = []
    counts: Dict[str, int] = {}
    parts: List[str] = []

    for session_id, group in sessions_dict.items():
        sessions.append(str(session_id))
        counts[str(session_id)] = len(group)
        lines = []

        for row in group:
            role = row.get("role", "")
            role_prefix = f"{role}: " if role else ""
            content = str(row.get(message_col, ""))
            lines.append(f"- {role_prefix}{content}")

        parts.append(f"#{session_id}\n" + "\n".join(lines))

    sessions_data = "\n\n".join(parts)
    return sessions, sessions_data, counts


def read_file_content(path: Path) -> str:
    """ファイルの拡張子に応じてテキストを抽出する"""
    suffix = path.suffix.lower()

    if suffix == ".txt":
        return path.read_text(encoding="utf-8", errors="replace")

    elif suffix == ".pdf":
        try:
            from pypdf import PdfReader
            reader = PdfReader(path)
            text = []
            for i, page in enumerate(reader.pages, start=1):
                page_text = page.extract_text() or ""
                # ページ番号マーカーを挿入
                text.append(f"[p.{i}]\n{page_text}")
            return "\n\n".join(text)
        except ImportError:
            return "[Error] pypdf not installed"
        except Exception as e:
            return f"[Error reading PDF] {e}"

    elif suffix == ".xlsx":
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, data_only=True)
            text = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    row_text = [str(cell) for cell in row if cell is not None]
                    if row_text:
                        text.append("\t".join(row_text))
            return "\n".join(text)
        except ImportError:
            return "[Error] openpyxl not installed"
        except Exception as e:
            return f"[Error reading Excel] {e}"

    elif suffix == ".csv":
        try:
            text = []
            with open(path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f)
                for row in reader:
                    text.append(",".join(row))
            return "\n".join(text)
        except Exception as e:
            return f"[Error reading CSV] {e}"

    elif suffix == ".docx":
        try:
            import docx
            doc = docx.Document(path)
            return "\n".join([para.text for para in doc.paragraphs])
        except ImportError:
            return "[Error] python-docx not installed"
        except Exception as e:
            return f"[Error reading DOCX] {e}"

    else:
        # その他のファイルはテキストとして試みるか、スキップ
        try:
            return path.read_text(encoding="utf-8", errors="replace")
        except:
            return f"[Unsupported file type: {suffix}]"


def load_documents_from_folder(folder_path: Path) -> str:
    """指定フォルダ内の全ファイルを読み込んで結合テキストを返す"""
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    combined_text = []
    # 隠しファイル以外を対象にする
    files = sorted([p for p in folder_path.iterdir() if p.is_file() and not p.name.startswith(".")])

    if not files:
        return "No files found in the directory."

    for file_path in files:
        content = read_file_content(file_path)
        combined_text.append(f"\n\n=== File: {file_path.name} ===\n\n{content}")

    return "\n".join(combined_text)


def iter_documents_from_folder(folder_path: Path) -> Iterator[Tuple[str, str]]:
    """指定フォルダ内のファイルを順次読み込んで (ファイル名, コンテンツ) をyieldする"""
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    # 隠しファイル以外を対象にする（metadata.jsonも除外）
    files = sorted([
        p for p in folder_path.iterdir()
        if p.is_file() and not p.name.startswith(".") and p.name != "metadata.json"
    ])

    for file_path in files:
        content = read_file_content(file_path)
        yield file_path.name, content


def iter_documents_with_metadata(folder_path: Path) -> Iterator[DocumentWithMetadata]:
    """
    指定フォルダ内のファイルをメタデータ付きで順次読み込む

    metadata.json が存在する場合、各ファイルにURL等のメタデータを付与する
    """
    if not folder_path.exists():
        raise FileNotFoundError(f"Folder not found: {folder_path}")

    # スクレイパーのメタデータを読み込み
    scraper_metadata = load_scraper_metadata(folder_path)

    # 隠しファイルとmetadata.json以外を対象にする
    files = sorted([
        p for p in folder_path.iterdir()
        if p.is_file() and not p.name.startswith(".") and p.name != "metadata.json"
    ])

    for file_path in files:
        content = read_file_content(file_path)
        filename = file_path.name

        # メタデータを取得（存在しない場合は空辞書）
        meta = scraper_metadata.get(filename, {})

        # 日付を抽出
        page_title = meta.get("page_title", "")
        link_text = meta.get("link_text", "")
        extracted_date = extract_date_from_metadata(page_title, link_text, filename)

        yield DocumentWithMetadata(
            filename=filename,
            content=content,
            url=meta.get("source_url", ""),
            page_title=page_title,
            link_text=link_text,
            page_url=meta.get("page_url", ""),
            date=extracted_date
        )


def build_document_registry(folder_path: Path) -> Tuple[List[DocumentWithMetadata], dict]:
    """
    フォルダ内の全ドキュメントを読み込み、引用レジストリ用の情報を構築

    Returns:
        (documents, registry_data)
        - documents: DocumentWithMetadata のリスト
        - registry_data: {filename: {url, page_title, ...}} の辞書
    """
    documents = list(iter_documents_with_metadata(folder_path))

    registry_data = {}
    for doc in documents:
        registry_data[doc.filename] = {
            "url": doc.url,
            "page_title": doc.page_title,
            "link_text": doc.link_text,
            "page_url": doc.page_url,
            "date": doc.date
        }

    return documents, registry_data

